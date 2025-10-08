from django.db.models import Prefetch
from django_filters.rest_framework import DjangoFilterBackend
from institution.models import Institution, InstitutionBranch
from order.feedback.models import DeliveryFeedback, InstitutionFeedback
from order.feedback.serializers import DeliveryFeedbackSerializer, InstitutionFeedbackSerializer
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from base.api_views import CustomPagination, MultiSerializerViewSetMixin, DynamicPagination
from crm.api.order.filters import OrderItemGroupFilter
from crm.api.order.serializers import (
    CrmOrderItemGroupSerializer,
    CrmOrderItemGroupDetailSerializer,
    CrmOrderSerializer,
    CrmOrderItemSerializer,
    CrmOrderItemUpdateSerializer,
    CrmOrderItemCreateSerializer,
    CrmOderUpdateSerializer,
)
from crm.api.order.services import recalculate_order_products, change_order_courier, add_order_item
from order.models import OrderItemGroup, OrderItem, Order
from order.status_controller import cancel_order, update_order_status
from .permissions import OrderPermission


class OrderViewSet(MultiSerializerViewSetMixin, ModelViewSet):
    queryset = Order.objects.select_related("customer", "courier", "operator")
    serializer_class = CrmOrderSerializer
    serializer_action_classes = {
        "list": CrmOrderSerializer,
        "retrieve": CrmOrderSerializer,
        "update_status": CrmOderUpdateSerializer,
        "change_courier": CrmOderUpdateSerializer,
    }
    permission_classes = [IsAuthenticated, OrderPermission]
    filter_backends = [DjangoFilterBackend]
    pagination_class = DynamicPagination

    def get_queryset(self):
        qs = super().get_queryset()

        if self.request.user.type in ["institution_admin", "institution_owner"]:
            institution = self.request.user.get_institution()
            if institution:
                qs = qs.filter(item_groups__institution=institution).distinct()
            else:
                qs = qs.none()
        return qs

    def update_status(self, request, *args, **kwargs):
        order = self.get_object()
        serializer = self.get_serializer(data=request.data)
        
        serializer.is_valid(raise_exception=True)
        status = serializer.validated_data.get("status")
        preparing_time = serializer.validated_data.get("preparing_time")
        
        if status and status != order.status:
            if status == "rejected":
                success, message = cancel_order(order, ignore_constraints=True)
                if success:
                    return Response({"message": message})
                else:
                    return Response({"message": message}, status=400)
            else:
                success, message = update_order_status(order, status, preparing_time)
                if success is True:
                    return Response({"message": "success"})
                elif success is False:
                    return Response({"message": message}, status=400)
        
        return Response({"message": "success"})

    def change_courier(self, request, *args, **kwargs):
        order = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        courier = serializer.validated_data.get("courier")
        is_success, message = change_order_courier(order, courier)
        if not is_success:
            return Response({"message": message}, status=400)
        return Response(serializer.data)


class OrderItemGroupViewSet(MultiSerializerViewSetMixin, ModelViewSet):
    queryset = OrderItemGroup.objects.select_related(
        "order", "order__customer", "order__courier", "order__operator"
    )
    serializer_action_classes = {
        "list": CrmOrderItemGroupSerializer,
        "retrieve": CrmOrderItemGroupDetailSerializer,
    }
    permission_classes = [IsAuthenticated, OrderPermission]
    filter_backends = [DjangoFilterBackend]
    pagination_class = DynamicPagination
    filterset_class = OrderItemGroupFilter

    def get_queryset(self):
        qs = super().get_queryset()

        if self.request.user.type in ["institution_admin", "institution_owner"]:
            institution = self.request.user.get_institution()
            if institution:
                qs = qs.filter(institution=institution)
            else:
                qs = qs.none()

        regions = self.request.query_params.get("regions")
        if regions:
            qs = qs.filter(institution_branch__address__region__name__icontains=regions)

        customer_phone_number = self.request.query_params.get("customer_phone_number")
        if customer_phone_number:
            qs = qs.filter(order__customer__phone_number__icontains=customer_phone_number)

        if self.action in ["retrieve"]:
            qs = qs.prefetch_related(
                Prefetch(
                    "items",
                    OrderItem.objects.select_related("product").prefetch_related("options__option"),
                )
            )
        return qs.order_by("-id")


class OrderItemViewSet(MultiSerializerViewSetMixin, ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_action_classes = {
        "list": CrmOrderItemSerializer,
        "retrieve": CrmOrderItemSerializer,
        "create": CrmOrderItemCreateSerializer,
        "partial_update": CrmOrderItemUpdateSerializer,
    }
    permission_classes = [IsAuthenticated, OrderPermission]
    filter_backends = [DjangoFilterBackend]
    pagination_class = DynamicPagination

    def partial_update(self, request, *args, **kwargs):
        order_item: OrderItem = self.get_object()
        serializer = self.get_serializer(order_item, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        recalculate_order_products(order_item.order_item_group.order)
        serializer = self.get_serializer(order_item)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):

        data = request.data

        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        incident_product = request.data.get("incident_product", None)
        data = serializer.validated_data

        is_success, response = add_order_item(
            order_group=data["order_item_group"],
            product=data["product"],
            count=data["count"],
            options=data["options"],
            is_incident=incident_product
        )

        if not is_success:
            return Response({"message": response}, status=400)

        serializer.instance = response
        return Response(serializer.data, status=201)

    def destroy(self, request, *args, **kwargs):
        order_item = self.get_object()
        order_item.delete()
        recalculate_order_products(order_item.order_item_group.order)
        return Response(status=204)
    
import io
import pandas as pd
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from order.models import Order
from reportlab.lib.pagesizes import letter

from django.db.models import Count, Q, Sum, F
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from rest_framework.viewsets import ReadOnlyModelViewSet
from rest_framework.filters import OrderingFilter


import re

def clean_sheet_name(name: str, default: str = "Report") -> str:
    if not name:
        return default
    cleaned = re.sub(r'[:\\/?*\[\]]+', '', name)
    cleaned = cleaned.strip()
    cleaned = cleaned[:31]
    return cleaned if cleaned else default

class ReportAPIView(APIView):
    
    def get(self, request):
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            payment = request.query_params.get('payment')
            export_type = request.query_params.get('export_type')
            institution = request.query_params.get('institution', request.user.get_institution())
            institution_branch = request.query_params.get('institution_branch')
            
            if start_date and end_date:
                start_date = timezone.make_aware(parse_datetime(start_date))
                end_date = timezone.make_aware(parse_datetime(end_date))
            else:
                now = timezone.now()
                start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                end_date = now
                
            orders = Order.objects.select_related('customer', 'operator').prefetch_related('item_groups__institution_branch__institution','item_groups__items__product'
            ).filter(created_at__gte=start_date, created_at__lte=end_date)
            
            
            if payment:
                orders = orders.filter(payment_method=payment)   

            if institution_branch:
                orders = orders.filter(item_groups__institution_branch__in=institution_branch.split(','))
            else:
                orders = orders.filter(item_groups__institution_branch__institution=institution)

            
            order_stats = orders.aggregate(
                closed_orders=Count('id', filter=Q(status='closed')),  # 'closed' statusidagi buyurtmalar soni
                rejected_orders=Count('id', filter=Q(status='rejected')),  # 'rejected' statusidagi buyurtmalar soni
            )
            orders = orders.filter(status='closed')
            
            detail =  {
                "products_sum": 0,
                "cash_sum": 0,
                "payme_sum": 0,
                "payme_commision": 1,
                "payme_commision_sum": 0,
                "commission": 0,
                "income": 0,
                "products_count": 0,
                "price_diff_sum": 0,
                "delivery_sum": 0,
                "delivery_discount_sum": 0,
                "order_count": order_stats['closed_orders'],
                "order_uncompleted_count": order_stats['rejected_orders'],
                "diff_sum": 0
            }
            
            # Ma'lumotlarni yig'ish
            
            rows = []
            for order in orders:
                for item_group in order.item_groups.all():
                    branch = item_group.institution_branch
                    if not branch or not branch.institution:
                        continue
                    
                    institution = branch.institution
                    tax_percentage = institution.tax_percentage_ordinary or 0
                    detail['commission'] = tax_percentage
                    
                    for item in item_group.items.all():
                        
                        product = item.product
                        options_sum = sum([option.adding_price for option in item.options.all()])
                        products_sum = product.price + options_sum
                        total_item_sum = products_sum * item.count
                        income = total_item_sum * (tax_percentage / 100)
                        
                        detail['income'] += round(income, 2)
                        detail['products_count'] += item.count
                        detail["products_sum"] += round(total_item_sum, 2)
                        detail["price_diff_sum"] += round((product.price - (product.old_price or 0)) * item.count, 2) if product.old_price else round(0, 2)
                        detail["delivery_sum"] += round(order.delivering_sum, 2)
                        
                        if order.payment_method == "cash": 
                            detail["cash_sum"] += round(total_item_sum, 2)
                        else:
                            detail["payme_sum"] += round(total_item_sum, 2)
                            

                        rows.append({
                            "ИД заказа": order.id,
                            "Дата": order.created_at.strftime("%Y-%m-%d"),
                            "Легальная название компании": institution.legal_name,
                            "Название товара": product.name,
                            "Склад": branch.name,
                            "Цена товара": round(products_sum, 2),
                            "Кол-во": item.count,
                            "Общая сумма товара": round(total_item_sum, 2),
                            "Цена скидки": 0,
                            "Доход": round(income, 2),
                            "Сум или процент": f"{tax_percentage}%",
                            "Тип оплаты": "Payme" if order.payment_method == "payme" else "Наличные",
                        })


            df = pd.DataFrame(rows)

            detail["payme_commision_sum"] = round(detail["payme_sum"] * (1/100), 2)
            detail["income"] = round(detail["products_sum"] * (detail["commission"]/100), 2)
            
            if detail["payme_sum"] == 0:
                detail["diff_sum"] = round(detail['income'], 2)
            else:
                detail['diff_sum'] = round((detail['products_sum'] - detail['cash_sum'] - detail['income']-detail["payme_commision_sum"]), 2)
            
            if detail['diff_sum'] < detail['payme_sum']:
                detail['diff_sum'] = -detail['diff_sum']

            if export_type == "xls":
                
                sheet_name = clean_sheet_name(institution.name)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]

                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font, Alignment

                    for col_idx, column in enumerate(df.columns, 1):
                        max_length = max(
                            df[column].astype(str).map(len).max(),
                            len(column)
                        )
                        adjusted_width = max_length + 4  # extra space
                        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                output.seek(0)
                filename = f"{sheet_name}_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response = HttpResponse(
                    output,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            else:
                return Response(detail, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class OrderStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.now()
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # orders = Order.objects.filter(status="closed", created_at__gte=start_of_month)
        orders = Order.objects.filter(
            status="closed",
            created_at__gte=start_of_month
        ).filter(
            Q(payment_method="payme", is_paid=True) | Q(payment_method="cash")
        )

        product_total = orders.aggregate(total=Sum("total_sum"))["total"] or 0
        cash_total = orders.filter(payment_method="cash").aggregate(total=Sum("total_sum"))["total"] or 0
        payme_total = orders.filter(payment_method="payme").aggregate(total=Sum("total_sum"))["total"] or 0
        order_ids = orders.values_list("id", flat=True)
        total_commission = OrderItemGroup.objects.filter(
            order_id__in=order_ids
        ).aggregate(total=Sum("commission"))["total"] or 0

        return Response({
            "product_total": round(product_total, 2),
            "cash_total": round(cash_total, 2),
            "payme_total": round(payme_total, 2),
            "total_commission": round(total_commission, 2),
        })

class InstitutionFeedbackViewSet(ReadOnlyModelViewSet):
    queryset = InstitutionFeedback.objects.select_related("order__customer", "institution")
    # queryset = InstitutionFeedback.objects.select_related("order__customer").prefetch_related("order__item_groups__institution", "order__item_groups__institution_branch")
    serializer_class = InstitutionFeedbackSerializer
    permission_classes = [IsAuthenticated, OrderPermission]  # optional
    pagination_class = CustomPagination
    
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ["value", "created"]
    ordering = ["-created"]


class CourierFeedbackViewSet(ReadOnlyModelViewSet):
    queryset = DeliveryFeedback.objects.select_related("order__customer", "courier").order_by("-created")
    serializer_class = DeliveryFeedbackSerializer
    permission_classes = [IsAuthenticated, OrderPermission]  # optional
    pagination_class = CustomPagination

    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ["value", "created"]
    ordering = ["-created"]
